from openpyxl import Workbook
import random

def Xexl(x,y,count):
    wb = Workbook()
    sheet1 = wb.active
    file_name = ''
    sheet1.cell(row = count,column = 1).value = x
    sheet1.cell(row = count,column = 2).value = y
    wb.save(filename = file_name)
    return count+1
